from itertools import chain
from typing import List
import json

import bs4
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from parcoursup import Wish, WishData, session

# setup session and logins
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36 OPR/98.0.0.0"
headers = {"User-Agent": USER_AGENT}

session.headers.update(headers)

URLs = {
    'login': "https://dossierappel.parcoursup.fr/Candidat/authentification"
}

with open("setup.json", "r") as f:
    data = json.load(f)
    assert "id" in data and "password" in data
    login = {"g_cn_cod": data.get("id"),
             "g_cn_mot_pas": data.get("password"),
             "ACTION": 1,
             "CSRFToken": ""}


# retrieves the csrf token and logs the session to parcoursup with given logins
login_page = session.get(URLs['login'])
login_soup = bs4.BeautifulSoup(login_page.content, features="html.parser")
csrf_token = login_soup.find("input", {'name': 'CSRFToken'})['value']
login['CSRFToken'] = csrf_token

r = session.post(URLs['login'], data=login, headers=headers)

# Makes the list of wishes that the page contains
soup = bs4.BeautifulSoup(r.content, features="html.parser")

wishes = [WishData.from_html(s) for s in soup.find_all(
    'div', {'class': 'psup-wish-card'})]

# sorts wishes together if they are with/without boarding and removes the abandoned ones
ordered_wishes = []
abandoned_wishes = []


for w in wishes:
    ordered_wishes.append([w])
    if "internat" in w.school_type.lower():
        i = w.school_type.lower().find("internat")
        new_type = w.school_type[:i-5]

        associated_wish = None
        for ass_w in wishes:
            if ass_w == w:
                continue
            if "internat" in ass_w.school_type and ass_w.school == w.school:
                if new_type in ass_w.school_type:
                    ordered_wishes[-1].append(ass_w)
                    wishes.remove(ass_w)
                    break

for i in range(len(ordered_wishes)):
    for j in range(len(ordered_wishes[i])-1, -1, -1):
        w = ordered_wishes[i][j]
        if ("renonc" in w.status.lower()) or ("refus" in w.status.lower()):
            abandoned_wishes.append(ordered_wishes[i].pop(j))

ordered_wishes = [w for w in ordered_wishes if len(w) != 0]


# For all non abandoned wishes, retrieves data about waiting queue and everything
wishes_data: List[Wish] = []

for w in ordered_wishes:

    if len(w) == 1:
        if "avec internat" in w[0].school_type:
            w = Wish.from_html(None, w[0].soup)
        else:
            w = Wish.from_html(w[0].soup)
    else:
        if "sans internat" in w[0].school_type:
            w = Wish.from_html(w[0].soup, w[1].soup)
        else:
            w = Wish.from_html(w[1].soup, w[0].soup)

    wishes_data.append(w)


wb = xl.Workbook()
ws = wb.active

# setup row 1
head = ["Formation", "Classement", "File d'attente", "Longueur FA", "Dernière proposition",
        "Dernier admis l'an dernier", "Internat", "Classement", "Places", "File d'attente"]
ws.append(head)


# write data in every row
for wish in wishes_data:
    w_data, q_data, b_data = wish.wishData, wish.queueData, wish.boardingData
    row = [w_data.full_name]
    if q_data:
        row += [q_data.ranking, q_data.queue_ranking,
                q_data.queue_len, q_data.last_sent, q_data.last_admitted]
    else:
        row += ["", "", "", "", ""]
    if b_data:
        row += ["Oui", b_data.ranking, b_data.places, b_data.queue_ranking]
    else:
        row += ["Non"]
    print(row)
    ws.append(row)

for wish in abandoned_wishes:
    ws.append([wish.full_name])

# style the cells
accepted = PatternFill(fgColor="3cff2e", fill_type="solid")
waiting = PatternFill(fgColor="5b7a80", fill_type="solid")
abandoned = PatternFill(fgColor="474747", fill_type="solid")

name_column = ws['A']
for i, wish in enumerate(wishes_data):
    cell = name_column[i+1]
    status = wish.wishData.status
    if "accept" in status.lower():
        cell.fill = accepted
    elif "attente" in status.lower():
        cell.fill = waiting

for i, wish in enumerate(abandoned_wishes):
    cell = name_column[i+len(wishes_data)+1]
    cell.fill = abandoned


ws.column_dimensions['A'].width = 60
for i in range(1, len(head)+1):
    ws.column_dimensions[get_column_letter(i+1)].width = 20

wb.save(filename="test.xlsx")